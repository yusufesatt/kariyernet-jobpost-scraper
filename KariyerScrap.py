from requests import get
from bs4 import BeautifulSoup as bs
import jpype
import pandas as pd
import openpyxl
import asposecells
import requests

try:
    url = input("Kariyer.net ilan kategori linki giriniz: ")
    response = bs(get(url).text, 'html.parser')
    ilan = response.find("div", attrs={"class" : "list-items-wrapper"})
    ilanlar = ilan.find_all("div", attrs={"class" : "list-items"})
    def tum_iller():
        return "Adana, Adıyaman, Afyon, Ağrı, Amasya, Ankara, Antalya, Artvin, Aydın, Balıkesir, Bilecik, Bingöl, Bitlis, Bolu, Burdur, Bursa, Çanakkale, Çankırı, Çorum, Denizli, Diyarbakır, Edirne, Elazığ, Erzincan, Erzurum, Eskişehir, Gaziantep, Giresun, Gümüşhane, Hakkari, Hatay, Isparta, Mersin, İstanbul(Avr.), İzmir, Kars, Kastamonu, Kayseri, Kırklareli, Kırşehir, Kocaeli, Konya, Kütahya, Malatya, Manisa, Kahramanmaraş, Mardin, Muğla, Muş, Nevşehir, Niğde, Ordu, Rize, Sakarya, Samsun, Siirt, Sinop, Sivas, Tekirdağ, Tokat, Trabzon, Tunceli, Şanlıurfa, Uşak, Van, Yozgat, Zonguldak, Aksaray, Bayburt, Karaman, Kırıkkale, Batman, Şırnak, Bartın, Ardahan, Iğdır, Yalova, Karabük, Kilis, Osmaniye, Düzce, İstanbul(Asya)"

    liste = []

    for linkler in ilanlar:
        linkend = linkler.a.get("href")
        linkstr = "https://www.kariyer.net"
        linkfns = linkstr + linkend
        response_2 = bs(get(linkfns).text, 'html.parser')
        basliklar = response_2.find("h1",attrs={"class":"title mb-0"})
        baslik = basliklar.find("span").text.strip()
        firma = basliklar.find("a").text.strip()
        lokasyon = response_2.find("div",attrs={"class" : "company-location"}).text.strip()
        if (lokasyon == tum_iller()): lokasyon = "Uzaktan Çalışabilir - Tüm İller"
        else:lokasyon = lokasyon
        calisma = response_2.find("div", attrs={"class" : "value"})
        isleyis = calisma.find("p" , attrs={"class" : "mb-0"}).text.strip()
        tecrube = response_2.find("span", attrs={"class" : "value-contain"}).text.strip()
        liste.append([linkfns, baslik, firma, lokasyon, isleyis,tecrube])

    liste = pd.DataFrame(liste, columns=(["İlan Url", "Başlık", "Firma", "Lokasyon", "İşleyiş", "Tecrübe"]))
    liste.to_excel("İlanlar.xlsx", sheet_name="ilan", index=False)

    jpype.startJVM()
    from asposecells.api import Workbook
    wb = Workbook("İlanlar.xlsx")
    worksheet = wb.getWorksheets().get(0)
    for i in range(6):
        worksheet.autoFitColumn(i)
    wb.save("İlanlar.xlsx")

    wb = openpyxl.load_workbook('İlanlar.xlsx')
    wb.sheetnames
    std=wb['Evaluation Warning']
    wb.remove(std)
    wb.save('İlanlar.xlsx')
    print("Başarıyla excel'e aktarıldı :)")
except:
    print("Url'yi kontrol ediniz!")


